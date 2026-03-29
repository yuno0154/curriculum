"""
streamlit_app.py — 성취기준 검색 & 편집 (Streamlit Cloud 버전)
편집 내용은 GitHub API를 통해 edits.json 파일로 저장됩니다.
"""

import streamlit as st
import json, re, requests, base64, io, os
from datetime import datetime
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────
#  1. 과목명 매핑
# ─────────────────────────────────────────────────────────
SUBJECT_MAP = {
    '10공국1':'공통국어1','10공국2':'공통국어2',
    '10공수1':'공통수학1','10공수2':'공통수학2',
    '10기수1':'기본수학1','10기수2':'기본수학2',
    '10공영1':'공통영어1','10공영2':'공통영어2',
    '10기영1':'기본영어1','10기영2':'기본영어2',
    '10통사1':'통합사회1','10통사2':'통합사회2',
    '10한사1':'한국사1','10한사2':'한국사2',
    '10통과1':'통합과학1','10통과2':'통합과학2',
    '10과탐1':'과학탐구실험1','10과탐2':'과학탐구실험2',
    '12화언':'화법과 언어','12독작':'독서와 작문','12문학':'문학',
    '12독토':'독서 토론과 글쓰기','12매의':'매체 의사소통',
    '12언탐':'언어생활 탐구','12직의':'직무 의사소통',
    '12문영':'문학과 영상','12주탐':'주제 탐구 독서',
    '12대수':'대수','12기하':'기하','12확통':'확률과 통계',
    '12수과':'수학과제 탐구','12경수':'경제 수학','12인수':'인공지능 수학',
    '12직수':'직무 수학','12실통':'실용 통계','12수문':'수학과 문화',
    '12영독':'영어 독해와 작문','12영발':'영어 발표와 토론',
    '12세영':'세계 문화와 영어','12미영':'미디어 영어',
    '12직영':'직무 영어','12실영':'실용 영어',
    '12심영':'심화 영어','12영문':'영어문학읽기',
    '12독어':'독일어','12독회':'독일어 회화','12독문':'독일어 독해와 작문','12심독':'심화 독일어',
    '12프어':'프랑스어','12프회':'프랑스어 회화','12프문':'프랑스어 독해와 작문','12심프':'심화 프랑스어',
    '12스어':'스페인어','12스회':'스페인어 회화','12심스':'심화 스페인어',
    '12중어':'중국어','12중회':'중국어 회화','12중문':'중국어 독해와 작문','12심중':'심화 중국어',
    '12일어':'일본어','12일회':'일본어 회화','12일문':'일본어 독해와 작문','12심일':'심화 일본어',
    '12러어':'러시아어','12러회':'러시아어 회화','12러문':'러시아어 독해와 작문','12심러':'심화 러시아어',
    '12아어':'아랍어','12아회':'아랍어 회화','12아문':'아랍어 독해와 작문','12심아':'심화 아랍어',
    '12베어':'베트남어','12베회':'베트남어 회화','12베문':'베트남어 독해와 작문','12심베':'심화 베트남어',
    '12한탐':'한국사 탐구','12동역':'동아시아 역사 기행','12세사':'세계사',
    '12역현':'역사로 탐구하는 현대 세계','12법사':'법과 사회','12정치':'정치',
    '12경제':'경제','12사문':'사회문제 탐구','12사탐':'사회과학 탐구',
    '12기지':'기후변화와 지속가능한 세계','12금융':'금융과 경제생활',
    '12국관':'국제 관계의 이해','12세지':'세계지리','12여지':'여행지리',
    '12도탐':'도시의 미래 탐구',
    '12물리':'물리학','12화학':'화학','12생과':'생명과학','12지구':'지구과학',
    '12역학':'역학과 에너지','12전자':'전자기와 양자','12물에':'물질과 에너지',
    '12반응':'화학 반응의 세계','12세포':'세포와 물질대사','12유전':'생물의 유전',
    '12행우':'행성우주과학','12지시':'지구시스템과학','12기환':'기후변화와 환경생태',
    '12과사':'과학의 역사와 문화','12융탐':'융합과학탐구',
    '12윤사':'윤리와 사상','12윤탐':'윤리 문제 탐구','12인윤':'인간과 윤리','12현윤':'현대사회와 윤리',
    '12기가':'기술·가정','12로봇':'로봇과 공학세계','12생활':'생활과학 탐구',
    '12아동':'아동발달과 부모','12자립':'생애 설계와 자립','12지재':'지식재산 일반','12창공':'창의 공학 설계',
    '12정':'정보','12데과':'데이터 과학','12소생':'소프트웨어와 생활','12인기':'인공지능 기초',
    '12음':'음악','12감비':'음악 감상과 비평','12연창':'음악 연주와 창작','12음미':'음악과 미디어',
    '12미':'미술','12미감':'미술 감상과 비평','12미매':'미술과 매체','12미창':'미술 창작',
    '12체육1':'체육1','12체육2':'체육2',
    '12스과':'스포츠 과학','12스생1':'스포츠 생활1','12스생2':'스포츠 생활2','12운건':'운동과 건강',
    '12연극':'연극',
    '12논리':'논리와 사고','12논술':'논술','12보건':'보건','12삶종':'삶과 종교',
    '12생환':'생태와 환경','12심리':'심리학','12인경':'인간과 경제활동',
    '12인철':'인간과 철학','12진로':'진로와 직업','12교이':'교육의 이해',
    '12한문':'한문','12한고':'한문 고전 읽기','12언한':'언어생활과 한자',
}
AREA_MAP = {
    '체육과': {'12스문': '스포츠 문화'},
    '영어과': {'12스문': '스페인어 독해와 작문'},
}

# ─────────────────────────────────────────────────────────
#  1-b. 분류 체계
# ─────────────────────────────────────────────────────────
LEVEL_ORDER = [
    '중학교', '고등학교',
    '선택', '생활 외국어',
    '과학 계열', '체육 계열', '예술 계열',
    '전문 공통',
    '경영·금융', '보건·복지', '문화·예술·디자인·방송', '미용',
    '관광·레저', '식품·조리', '건축·토목', '기계', '재료',
    '화학 공업', '섬유·의류', '전기·전자', '정보·통신',
    '환경·안전·소방', '농림·축산', '수산·해운', '융복합·지식 재산',
]

# 화면 표시용 이름 (내부 level key → 표시 문자열)
LEVEL_DISPLAY = {
    '생활 외국어': '(선택) 생활 외국어',
}

LEVEL_GROUP = {
    '중학교': '학교 교육과정', '고등학교': '학교 교육과정',
    '선택': '학교 교육과정', '생활 외국어': '학교 교육과정',
    '과학 계열': '계열 선택 과목', '체육 계열': '계열 선택 과목', '예술 계열': '계열 선택 과목',
}
for _lv in LEVEL_ORDER[7:]:
    LEVEL_GROUP[_lv] = '전문교과'

GROUP_ORDER = ['학교 교육과정', '계열 선택 과목', '전문교과']

def get_code_prefix(code):
    prefix = code.strip('[]').split('-')[0]
    return prefix[:-2] if re.match(r'^12[가-힣A-Za-z]+0[1-9]$', prefix) else prefix

def get_subject_name(area, code):
    p = get_code_prefix(code)
    return AREA_MAP.get(area, {}).get(p) or SUBJECT_MAP.get(p) or p

# ─────────────────────────────────────────────────────────
#  2. 데이터 로드 & 인덱스 구축
#  data_size를 캐시 키로 사용 — data.json 변경 시 자동 갱신
# ─────────────────────────────────────────────────────────
@st.cache_data(show_spinner='데이터 로드 중...')
def build_index(data_size: int = 0):  # data_size는 캐시 키 전용 — _ 접두사 금지
    """data.json을 읽고 렌더링용 인덱스를 구축 (캐시됨)."""
    with open('data.json', encoding='utf-8') as f:
        raw = json.load(f)

    by_level: dict[str, list] = {}
    gc: dict[str, int] = {g: 0 for g in GROUP_ORDER}

    for d in raw:
        lv = d['level']
        by_level.setdefault(lv, []).append(d)
        grp = LEVEL_GROUP.get(lv)
        if grp in gc:
            gc[grp] += 1

    for lv in by_level:
        by_level[lv].sort(key=lambda x: x['code'])

    # 고등학교: area → subject_name → [items]
    hs_tree: dict[str, dict[str, list]] = {}
    for d in by_level.get('고등학교', []):
        area = d['subject']
        sn = get_subject_name(area, d['code'])
        hs_tree.setdefault(area, {}).setdefault(sn, []).append(d)

    lv_counts = {lv: len(items) for lv, items in by_level.items()}

    return raw, gc, by_level, hs_tree, lv_counts

def _gh_headers():
    token = st.secrets.get('GITHUB_TOKEN', '')
    return {'Authorization': f'token {token}', 'Accept': 'application/vnd.github.v3+json'} if token else {}

def _gh_repo():
    return st.secrets.get('GITHUB_REPO', 'yuno0154/curriculum')

@st.cache_data(ttl=60)
def load_edits_remote():
    """GitHub API에서 edits.json 읽기 (60초 캐시)"""
    try:
        url = f'https://api.github.com/repos/{_gh_repo()}/contents/edits.json'
        r = requests.get(url, headers=_gh_headers(), timeout=5)
        if r.ok:
            raw = base64.b64decode(r.json()['content']).decode('utf-8')
            return json.loads(raw)
    except Exception:
        pass
    # 로컬 폴백
    try:
        with open('edits.json', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {}

def save_edits_remote(edits: dict, commit_msg: str) -> tuple[bool, str]:
    """GitHub API로 edits.json 커밋"""
    try:
        token = st.secrets.get('GITHUB_TOKEN', '')
        if not token:
            return False, 'GitHub 토큰이 설정되지 않았습니다. (Streamlit Secrets 확인)'
        url = f'https://api.github.com/repos/{_gh_repo()}/contents/edits.json'
        headers = _gh_headers()

        # 현재 SHA 조회 (업데이트 시 필요)
        r = requests.get(url, headers=headers, timeout=5)
        sha = r.json().get('sha') if r.ok else None

        content = base64.b64encode(
            json.dumps(edits, ensure_ascii=False, indent=2).encode('utf-8')
        ).decode()

        payload = {
            'message': commit_msg,
            'content': content,
            'committer': {'name': '성취기준 시스템', 'email': 'noreply@school.kr'},
        }
        if sha:
            payload['sha'] = sha

        r = requests.put(url, headers=headers, json=payload, timeout=10)
        if r.ok:
            load_edits_remote.clear()   # 캐시 무효화
            return True, 'GitHub에 저장됐습니다.'
        else:
            return False, r.json().get('message', '저장 실패')
    except Exception as e:
        return False, str(e)

# ─────────────────────────────────────────────────────────
#  3. 페이지 설정
# ─────────────────────────────────────────────────────────
st.set_page_config(
    page_title='성취기준 검색 | 사곡고등학교',
    page_icon='📚',
    layout='wide',
    initial_sidebar_state='expanded',
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Pretendard:wght@400;500;600;700;800;900&display=swap');

/* ── 전체 배경 & 폰트 ─────────────────────────────── */
html, body, .stApp {
    background: #f4f6fb !important;
    font-family: 'Pretendard', -apple-system, sans-serif !important;
    color: #1e293b !important;
}
.block-container { padding-top: 1.2rem !important; max-width: 1100px !important; }

/* 사이드바 */
[data-testid="stSidebar"] {
    background: #ffffff !important;
    border-right: 1px solid #e2e8f0 !important;
    display: block !important;
    visibility: visible !important;
    min-width: 240px !important;
}
[data-testid="stSidebar"] * { color: #334155 !important; }
[data-testid="stSidebarCollapsedControl"],
[data-testid="collapsedControl"] {
    display: flex !important;
    visibility: visible !important;
}

/* ── 헤더 카드 ─────────────────────────────────────── */
.hero-wrap {
    background: #ffffff;
    border-radius: 20px;
    padding: 1.6rem 2.5rem 1.4rem;
    margin-bottom: 1.8rem;
    border: 1px solid #e2e8f0;
    border-top: 4px solid #6366f1;
    box-shadow: 0 4px 16px rgba(0,0,0,0.06);
}
.hero-header {
    display: flex;
    align-items: center;
    justify-content: space-between;
    margin-bottom: 0.5rem;
}
.hero-title {
    font-size: 1.7rem;
    font-weight: 900;
    letter-spacing: -0.03em;
    background: linear-gradient(135deg, #4f46e5, #7c3aed);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    line-height: 1.25;
    margin: 0;
}
.hero-sub {
    color: #64748b;
    font-size: 0.92rem;
    font-weight: 500;
    margin-bottom: 0.75rem;
}
.hero-author {
    display: inline-block;
    background: #eef2ff;
    border: 1px solid #c7d2fe;
    color: #4338ca;
    font-size: 0.82rem;
    font-weight: 700;
    padding: 0.3rem 1rem;
    border-radius: 20px;
    letter-spacing: 0.02em;
    white-space: nowrap;
    flex-shrink: 0;
}
.hero-stats {
    display: flex;
    justify-content: center;
    gap: 0.75rem;
    margin-top: 1rem;
    flex-wrap: wrap;
}
.stat-item {
    background: #f8fafc;
    border: 1px solid #e2e8f0;
    border-radius: 10px;
    padding: 0.4rem 1.1rem;
    font-size: 0.83rem;
    color: #475569;
    font-weight: 500;
}
.stat-item b { color: #4f46e5; font-weight: 800; }

/* ── 탭 ────────────────────────────────────────────── */
div[data-baseweb="tab-list"] {
    background: #ffffff;
    border-radius: 12px;
    padding: 0.25rem;
    gap: 0.2rem;
    box-shadow: 0 1px 4px rgba(0,0,0,0.08);
    border: 1px solid #e2e8f0;
}
div[data-baseweb="tab"] {
    border-radius: 9px;
    font-weight: 600;
    font-size: 0.92rem;
    color: #64748b !important;
}
div[data-baseweb="tab"][aria-selected="true"] {
    background: #6366f1 !important;
    color: #ffffff !important;
}
button[data-baseweb="tab"]:hover { background: #f1f5f9 !important; }

/* ── 성취기준 코드 배지 ──────────────────────────────── */
.code-cell {
    font-family: 'Courier New', monospace;
    font-weight: 700;
    color: #4f46e5;
    font-size: 0.8rem;
    background: #eef2ff;
    border: 1px solid #c7d2fe;
    padding: 0.22rem 0.55rem;
    border-radius: 6px;
    display: inline-block;
    white-space: nowrap;
    line-height: 1.4;
}

/* ── 성취기준 진술문 ─────────────────────────────────── */
.stmt-cell {
    font-size: 0.96rem;
    line-height: 1.8;
    color: #1e293b;
    font-weight: 400;
}

/* ── 구분선 ─────────────────────────────────────────── */
hr { border-color: #e2e8f0 !important; margin: 0.5rem 0 !important; }

/* ── 섹션 제목 ──────────────────────────────────────── */
h3 { color: #1e293b !important; font-weight: 800 !important; letter-spacing: -0.02em; }
h4 { color: #374151 !important; font-weight: 700 !important; }

/* ── selectbox / radio / input ──────────────────────── */
div[data-testid="stSelectbox"] > div > div,
div[data-testid="stTextInput"] input {
    background: #ffffff !important;
    border: 1.5px solid #cbd5e1 !important;
    border-radius: 10px !important;
    color: #1e293b !important;
    font-size: 0.93rem !important;
}
div[data-testid="stSelectbox"] > div > div:focus-within,
div[data-testid="stTextInput"] input:focus {
    border-color: #6366f1 !important;
    box-shadow: 0 0 0 3px rgba(99,102,241,0.12) !important;
}

/* ── 버튼 ────────────────────────────────────────────── */
div[data-testid="stDownloadButton"] button,
div[data-testid="stButton"] button {
    border-radius: 10px !important;
    font-weight: 600 !important;
    font-size: 0.88rem !important;
    border: none !important;
    transition: all 0.2s !important;
}
div[data-testid="stDownloadButton"] button {
    background: linear-gradient(135deg, #6366f1, #8b5cf6) !important;
    color: #ffffff !important;
    box-shadow: 0 2px 8px rgba(99,102,241,0.3) !important;
}
div[data-testid="stDownloadButton"] button:hover {
    box-shadow: 0 4px 16px rgba(99,102,241,0.4) !important;
    transform: translateY(-1px) !important;
}

/* ── 성공/정보 메시지 ────────────────────────────────── */
div[data-testid="stSuccess"] { border-radius: 10px !important; }
div[data-testid="stInfo"]    { border-radius: 10px !important; }

/* ── expander ────────────────────────────────────────── */
div[data-testid="stExpander"] {
    background: #ffffff;
    border: 1px solid #e2e8f0 !important;
    border-radius: 12px !important;
    box-shadow: 0 1px 4px rgba(0,0,0,0.05);
}
div[data-testid="stExpander"] summary {
    font-size: 0.9rem !important;
    font-weight: 600 !important;
    color: #374151 !important;
}

/* ── metric ─────────────────────────────────────────── */
div[data-testid="stMetric"] {
    background: #ffffff;
    border: 1px solid #e2e8f0;
    border-radius: 12px;
    padding: 0.75rem 1rem;
    box-shadow: 0 1px 4px rgba(0,0,0,0.05);
}
div[data-testid="stMetricValue"] { color: #4f46e5 !important; font-weight: 800 !important; }
div[data-testid="stMetricLabel"] { color: #64748b !important; font-weight: 600 !important; }

/* ── 다운로드 탭 카드 ────────────────────────────────── */
.dl-section {
    background: #ffffff;
    border: 1px solid #e2e8f0;
    border-radius: 16px;
    padding: 1.5rem;
    margin-bottom: 1rem;
    box-shadow: 0 2px 8px rgba(0,0,0,0.05);
}
.dl-section-title {
    font-size: 1.05rem;
    font-weight: 800;
    color: #1e293b;
    margin-bottom: 1rem;
    display: flex;
    align-items: center;
    gap: 0.5rem;
}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────
#  4. 세션 상태 초기화
# ─────────────────────────────────────────────────────────
if 'edits' not in st.session_state:
    st.session_state.edits = load_edits_remote()

_data_size = os.path.getsize('data.json')
data, _gc, _by_level, _hs_tree, _lv_counts = build_index(_data_size)

def get_stmt(item):
    return st.session_state.edits.get(item['code'], {}).get('statement', item['statement'])

def is_edited(code):
    return code in st.session_state.edits

# ─────────────────────────────────────────────────────────
#  5. 헤더
# ─────────────────────────────────────────────────────────
edited_cnt = len(st.session_state.edits)

# _gc, _by_level, _hs_tree, _lv_counts 는 build_index()에서 사전 계산됨

st.markdown(f"""
<div class="hero-wrap">
    <div class="hero-header">
        <div class="hero-title">📚 2022 개정 교육과정 성취기준 통합검색 시스템</div>
        <div class="hero-author">사곡고등학교 수석교사 최연호</div>
    </div>
    <div class="hero-sub">성취기준 검색 · 과목별 탐색 · 오류 수정</div>
    <div class="hero-stats">
        <div class="stat-item">학교 교육과정 <b>{_gc['학교 교육과정']:,}</b>건</div>
        <div class="stat-item">계열 선택 과목 <b>{_gc['계열 선택 과목']:,}</b>건</div>
        <div class="stat-item">전문교과 <b>{_gc['전문교과']:,}</b>건</div>
        <div class="stat-item">전체 <b>{len(data):,}</b>건</div>
        {'<div class="stat-item">수정됨 <b style=\'color:#f472b6\'>' + str(edited_cnt) + '</b>건</div>' if edited_cnt else ''}
    </div>
</div>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────
#  6. 엑셀 생성 함수
# ─────────────────────────────────────────────────────────
def build_excel(items: list, sheet_name: str = '성취기준') -> bytes:
    """items 목록 → 스타일 적용 엑셀 bytes"""
    rows = []
    for d in sorted(items, key=lambda x: x['code']):
        if d['level'] == '고등학교':
            area = d['subject']
            sn   = get_subject_name(area, d['code'])
        else:
            area = ''
            sn   = d['subject']
        rows.append({
            '분류':         d['level'],
            '교과':         area,
            '과목':         sn,
            '성취기준 코드': d['code'],
            '성취기준':     get_stmt(d),
        })
    df = pd.DataFrame(rows)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name[:31])
        ws = writer.sheets[sheet_name[:31]]

        # 열 너비
        col_widths = [10, 14, 22, 20, 80]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # 헤더 스타일
        hdr_fill = PatternFill('solid', fgColor='4F46E5')
        hdr_font = Font(color='FFFFFF', bold=True, size=11)
        hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin = Side(style='thin', color='CCCCCC')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for cell in ws[1]:
            cell.fill   = hdr_fill
            cell.font   = hdr_font
            cell.alignment = hdr_align
            cell.border = border
        ws.row_dimensions[1].height = 22

        # 데이터 행 스타일
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            bg = 'F8F9FF' if row_idx % 2 == 0 else 'FFFFFF'
            fill = PatternFill('solid', fgColor=bg)
            for cell in row:
                cell.fill   = fill
                cell.border = border
                cell.alignment = Alignment(vertical='top', wrap_text=(cell.column == 5))

        ws.freeze_panes = 'A2'

    return buf.getvalue()

def dl_btn(label: str, items: list, filename: str, key: str):
    """다운로드 버튼 (데이터 없으면 disabled)"""
    if not items:
        st.caption('데이터 없음')
        return
    data_bytes = build_excel(items, filename[:31])
    st.download_button(
        label=label,
        data=data_bytes,
        file_name=f'{filename}.xlsx',
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        key=key,
        use_container_width=True,
    )

# ─────────────────────────────────────────────────────────
#  7. 메인 탭
# ─────────────────────────────────────────────────────────
tab_subject, tab_keyword, tab_download = st.tabs([
    '📂  과목별 탐색', '🔍  검색', '📥  엑셀 다운로드'
])

# ════════════════════════════════════════════════════════
#  탭 1: 과목별 탐색
# ════════════════════════════════════════════════════════
with tab_subject:
    # ── 분류 선택 (인덱스 사용 — O(1) 조회)
    ordered_levels = [l for l in LEVEL_ORDER if l in _by_level]

    col_grp, col_lv, col_info = st.columns([2, 2, 1])
    with col_grp:
        sel_group = st.selectbox(
            '교육과정 구분', GROUP_ORDER, key='tab1_group',
            format_func=lambda g: f'{g}  ({_gc.get(g, 0):,}건)'
        )
    with col_lv:
        group_levels = [l for l in ordered_levels if LEVEL_GROUP.get(l) == sel_group]
        level = st.selectbox(
            '분류', group_levels, key='tab1_level',
            format_func=lambda l: f'{LEVEL_DISPLAY.get(l, l)}  ({_lv_counts.get(l, 0):,}건)'
        )

    items_to_show = []
    title = ''

    if level == '고등학교':
        hs_area_tree = _hs_tree              # area → {sub_name → [items]}
        areas = sorted(hs_area_tree.keys())
        area_counts = {a: sum(len(v) for v in hs_area_tree[a].values()) for a in areas}
        col_a, col_b = st.columns(2)
        with col_a:
            area = st.selectbox(
                '교과', areas, key='hs_area',
                format_func=lambda a: f'{a}  ({area_counts[a]}개)'
            )
        with col_b:
            sub_dict = hs_area_tree.get(area, {})
            sub_names = sorted(sub_dict.keys())
            sub_counts = {sn: len(sub_dict[sn]) for sn in sub_names}
            sub_name = st.selectbox(
                '과목', sub_names, key=f'hs_sub_{area}',
                format_func=lambda s: f'{s}  ({sub_counts[s]}개)'
            )
        items_to_show = sub_dict.get(sub_name, [])
        title = f'{area} › {sub_name}'

    else:
        lv_items = _by_level.get(level, [])
        by_subject: dict[str, list] = {}
        for d in lv_items:
            by_subject.setdefault(d['subject'], []).append(d)
        subjects = sorted(by_subject.keys())
        subject = st.selectbox(
            '과목', subjects, key=f'other_subject_{level}',
            format_func=lambda s: f'{s}  ({len(by_subject[s])}개)'
        )
        items_to_show = by_subject.get(subject, [])
        title = f'{LEVEL_DISPLAY.get(level, level)} › {subject}'

    if items_to_show:
        with col_info:
            st.metric('성취기준', f'{len(items_to_show)}개')

        st.markdown(f'#### {title}')

        for item in items_to_show:
            stmt   = get_stmt(item)
            edited = is_edited(item['code'])

            c_code, c_stmt, c_edit = st.columns([2, 8, 1])

            with c_code:
                mark = ' 🔵' if edited else ''
                st.markdown(
                    f'<div class="code-cell">{item["code"]}{mark}</div>',
                    unsafe_allow_html=True
                )

            with c_stmt:
                st.markdown(f'<div class="stmt-cell">{stmt}</div>', unsafe_allow_html=True)

            with c_edit:
                if st.button('✏️', key=f'ebtn_{item["code"]}', help='수정'):
                    st.session_state[f'edit_open_{item["code"]}'] = True

            # 편집 패널
            if st.session_state.get(f'edit_open_{item["code"]}'):
                with st.container(border=True):
                    new_val = st.text_area(
                        f'진술문 수정 — {item["code"]}',
                        value=stmt,
                        key=f'ta_{item["code"]}',
                        height=90,
                    )
                    bc1, bc2, bc3 = st.columns([1, 1, 4])
                    with bc1:
                        if st.button('저장', key=f'save_{item["code"]}', type='primary'):
                            new_val = new_val.strip()
                            if new_val and new_val != item['statement']:
                                st.session_state.edits[item['code']] = {
                                    'statement': new_val,
                                    'original':  item['statement'],
                                    'edited_at': datetime.now().isoformat(timespec='seconds'),
                                }
                            else:
                                st.session_state.edits.pop(item['code'], None)

                            ok, msg = save_edits_remote(
                                st.session_state.edits,
                                f'성취기준 수정: {item["code"]}'
                            )
                            st.session_state[f'edit_open_{item["code"]}'] = False
                            if ok:
                                st.success(msg)
                            else:
                                st.warning(f'로컬 저장은 됐으나 GitHub 연동 실패: {msg}')
                            st.rerun()

                    with bc2:
                        if st.button('취소', key=f'cancel_{item["code"]}'):
                            st.session_state[f'edit_open_{item["code"]}'] = False
                            st.rerun()

                    if edited:
                        with bc3:
                            orig = st.session_state.edits[item['code']].get('original', item['statement'])
                            st.caption(f'원본: {orig}')
                        if st.button('↩ 원본 복원', key=f'reset_{item["code"]}'):
                            st.session_state.edits.pop(item['code'], None)
                            ok, msg = save_edits_remote(
                                st.session_state.edits,
                                f'성취기준 복원: {item["code"]}'
                            )
                            st.session_state[f'edit_open_{item["code"]}'] = False
                            st.rerun()

            st.divider()

# ════════════════════════════════════════════════════════
#  탭 2: 성취기준 키워드 검색
# ════════════════════════════════════════════════════════
with tab_keyword:
    search_mode = st.radio(
        '검색 유형',
        ['성취기준 검색', '과목명 검색'],
        horizontal=True,
        key='search_mode',
        label_visibility='collapsed',
    )

    # ── 공통 검색창 ──────────────────────────────────────
    if search_mode == '성취기준 검색':
        placeholder = '성취기준 키워드를 입력하세요 (예: 추론, 통계, 안전)'
    else:
        placeholder = '과목명을 입력하세요 (예: 수학, 미용, 전자, 체육)'

    query = st.text_input(
        '검색어',
        placeholder=placeholder,
        key='kw_input',
        label_visibility='collapsed',
    )

    if not query:
        if search_mode == '성취기준 검색':
            st.info('성취기준 키워드를 입력하면 전체 과목에서 검색합니다.')
        else:
            st.info('과목명 키워드를 입력하면 일치하는 과목 목록이 표시됩니다.')

    elif search_mode == '성취기준 검색':
        # ── 성취기준 키워드 검색 (기존 기능) ─────────────
        q = query.lower()
        filtered = sorted(
            [d for d in data
             if q in get_stmt(d).lower() or q in d['code'].lower()],
            key=lambda x: x['code']
        )
        st.caption(f'총 **{len(filtered)}**개 성취기준 검색됨')

        groups: dict[str, list] = {}
        for d in filtered:
            if d['level'] == '고등학교':
                sn  = get_subject_name(d['subject'], d['code'])
                grp = f'[고등학교] {d["subject"]} › {sn}'
            else:
                grp = f'[{LEVEL_DISPLAY.get(d["level"], d["level"])}] {d["subject"]}'
            groups.setdefault(grp, []).append(d)

        for grp_key, items in groups.items():
            with st.expander(f'**{grp_key}** — {len(items)}개', expanded=True):
                for item in items:
                    stmt = get_stmt(item)
                    highlighted = re.sub(
                        f'({re.escape(query)})', r'**\1**', stmt, flags=re.IGNORECASE
                    )
                    c1, c2 = st.columns([2, 8])
                    with c1:
                        mark = ' 🔵' if is_edited(item['code']) else ''
                        st.markdown(
                            f'<div class="code-cell">{item["code"]}{mark}</div>',
                            unsafe_allow_html=True
                        )
                    with c2:
                        st.markdown(highlighted)

    else:
        # ── 과목명 검색 ───────────────────────────────────
        q = query.lower()

        # 과목명 목록 구성 (level, subject_key, display_name)
        seen_subjects: dict[tuple, str] = {}  # (level, subject_key) → display_name
        for d in data:
            lv = d['level']
            sk = d['subject']
            if lv == '고등학교':
                sn = get_subject_name(sk, d['code'])
                seen_subjects.setdefault((lv, sk, sn), sn)
            else:
                seen_subjects.setdefault((lv, sk, sk), sk)

        # 검색어 매칭
        matched: list[tuple] = []
        for (lv, sk, sn) in seen_subjects:
            if q in sn.lower() or q in sk.lower():
                cnt = sum(
                    1 for d in data
                    if d['level'] == lv and d['subject'] == sk
                    and (lv != '고등학교' or get_subject_name(sk, d['code']) == sn)
                )
                matched.append((lv, sk, sn, cnt))

        # level 순서대로 정렬
        lv_idx = {l: i for i, l in enumerate(LEVEL_ORDER)}
        matched.sort(key=lambda x: (lv_idx.get(x[0], 99), x[2]))

        st.caption(f'**{len(matched)}**개 과목 검색됨')

        if not matched:
            st.warning('일치하는 과목이 없습니다.')
        else:
            for lv, sk, sn, cnt in matched:
                label = f'[{LEVEL_DISPLAY.get(lv, lv)}]  {sn}  —  {cnt}개'
                with st.expander(label, expanded=False):
                    # 해당 과목의 성취기준 목록
                    if lv == '고등학교':
                        sub_items = sorted(
                            [d for d in data
                             if d['level'] == lv and d['subject'] == sk
                             and get_subject_name(sk, d['code']) == sn],
                            key=lambda x: x['code']
                        )
                    else:
                        sub_items = sorted(
                            [d for d in data if d['level'] == lv and d['subject'] == sk],
                            key=lambda x: x['code']
                        )

                    # 다운로드 버튼
                    dl_btn(
                        f'⬇ {sn} 다운로드 ({cnt}개)',
                        sub_items,
                        f'{lv}_{sn}_성취기준',
                        f'dl_subsch_{lv}_{sk}_{sn}',
                    )
                    st.divider()

                    for item in sub_items:
                        stmt = get_stmt(item)
                        hi = re.sub(f'({re.escape(query)})', r'**\1**',
                                    sn, flags=re.IGNORECASE)
                        c1, c2 = st.columns([2, 8])
                        with c1:
                            mark = ' 🔵' if is_edited(item['code']) else ''
                            st.markdown(
                                f'<div class="code-cell">{item["code"]}{mark}</div>',
                                unsafe_allow_html=True
                            )
                        with c2:
                            st.markdown(
                                f'<div class="stmt-cell">{stmt}</div>',
                                unsafe_allow_html=True
                            )

# ════════════════════════════════════════════════════════
#  탭 3: 엑셀 다운로드
# ════════════════════════════════════════════════════════
with tab_download:
    st.markdown("""
    <div style="background:#eef2ff;border-left:4px solid #6366f1;border-radius:0 10px 10px 0;
                padding:0.8rem 1.2rem;margin-bottom:1.5rem;color:#3730a3;font-weight:600;font-size:0.9rem;">
        📌 수정된 성취기준이 있으면 수정본이 자동 반영됩니다.
    </div>""", unsafe_allow_html=True)

    dl_tab_school, dl_tab_track, dl_tab_voc = st.tabs(
        ['학교 교육과정', '계열 선택 과목', '전문교과']
    )

    # ── 학교 교육과정 (인덱스 사용) ─────────────────────────
    with dl_tab_school:
        mid_data  = _by_level.get('중학교', [])
        high_data = _by_level.get('고등학교', [])
        sel_data  = _by_level.get('선택', [])
        lofl_data = _by_level.get('생활 외국어', [])

        def _sub_index(items):
            idx: dict[str, list] = {}
            for d in items:
                idx.setdefault(d['subject'], []).append(d)
            return idx

        st.markdown('**중학교**')
        with st.container(border=True):
            col_m1, col_m2 = st.columns([1, 2])
            with col_m1:
                st.caption(f'전체 {len(mid_data)}개')
                dl_btn('⬇ 중학교 전체', mid_data, '중학교_성취기준_전체', 'dl_mid_all')
            with col_m2:
                mid_idx = _sub_index(mid_data)
                ms_sel  = st.selectbox('과목 선택', sorted(mid_idx), key='dl_ms_sel',
                                       format_func=lambda s: f'{s}  ({len(mid_idx[s])}개)')
                dl_btn(f'⬇ {ms_sel} ({len(mid_idx.get(ms_sel,[]))}개)',
                       mid_idx.get(ms_sel, []), f'중학교_{ms_sel}_성취기준', 'dl_mid_sub')

        if sel_data:
            st.markdown('**선택**')
            with st.container(border=True):
                c1, c2 = st.columns([1, 2])
                with c1:
                    st.caption(f'전체 {len(sel_data)}개')
                    dl_btn('⬇ 선택 전체', sel_data, '선택_성취기준_전체', 'dl_sel_all')
                with c2:
                    sel_idx = _sub_index(sel_data)
                    sel_sub_sel = st.selectbox('과목', sorted(sel_idx), key='dl_sel_sub_sel',
                                               format_func=lambda s: f'{s}  ({len(sel_idx[s])}개)')
                    dl_btn(f'⬇ {sel_sub_sel} ({len(sel_idx.get(sel_sub_sel,[]))}개)',
                           sel_idx.get(sel_sub_sel, []), f'선택_{sel_sub_sel}_성취기준', 'dl_sel_sub')

        if lofl_data:
            st.markdown('**(선택) 생활 외국어**')
            with st.container(border=True):
                c1, c2 = st.columns([1, 2])
                with c1:
                    st.caption(f'전체 {len(lofl_data)}개')
                    dl_btn('⬇ 생활 외국어 전체', lofl_data, '생활외국어_성취기준_전체', 'dl_lofl_all')
                with c2:
                    lofl_idx = _sub_index(lofl_data)
                    lofl_sel = st.selectbox('과목', sorted(lofl_idx), key='dl_lofl_sub_sel',
                                            format_func=lambda s: f'{s}  ({len(lofl_idx[s])}개)')
                    dl_btn(f'⬇ {lofl_sel} ({len(lofl_idx.get(lofl_sel,[]))}개)',
                           lofl_idx.get(lofl_sel, []), f'생활외국어_{lofl_sel}_성취기준', 'dl_lofl_sub')

        st.markdown('**고등학교**')
        with st.container(border=True):
            col_h1, col_h2, col_h3 = st.columns(3)
            hs_areas = sorted(_hs_tree.keys())
            area_cnt = {a: sum(len(v) for v in _hs_tree[a].values()) for a in hs_areas}
            with col_h1:
                st.caption(f'전체 {len(high_data)}개')
                dl_btn('⬇ 고등학교 전체', high_data, '고등학교_성취기준_전체', 'dl_high_all')
            with col_h2:
                st.caption('교과별')
                hs_area_sel = st.selectbox('교과 선택', hs_areas, key='dl_area_sel',
                                           format_func=lambda a: f'{a}  ({area_cnt[a]}개)')
                area_items = [d for sn_items in _hs_tree[hs_area_sel].values() for d in sn_items]
                dl_btn(f'⬇ {hs_area_sel} ({len(area_items)}개)', area_items,
                       f'고등학교_{hs_area_sel}_성취기준', 'dl_high_area')
            with col_h3:
                st.caption('과목별')
                hs_area2 = st.selectbox('교과', hs_areas, key='dl_area2_sel',
                                        format_func=lambda a: f'{a}  ({area_cnt[a]}개)')
                sub_dict2  = _hs_tree.get(hs_area2, {})
                sub_names2 = sorted(sub_dict2.keys())
                hs_sub_sel = st.selectbox('과목', sub_names2, key='dl_sub_sel',
                                          format_func=lambda s: f'{s}  ({len(sub_dict2[s])}개)')
                dl_btn(f'⬇ {hs_sub_sel} ({len(sub_dict2.get(hs_sub_sel,[]))}개)',
                       sub_dict2.get(hs_sub_sel, []),
                       f'고등학교_{hs_area2}_{hs_sub_sel}_성취기준', 'dl_high_sub')

    # ── 계열 선택 과목 (인덱스 사용) ─────────────────────────
    with dl_tab_track:
        track_levels = [l for l in ['과학 계열', '체육 계열', '예술 계열'] if l in _by_level]
        col_t1, col_t2 = st.columns([1, 2])
        with col_t1:
            track_all = [d for lv in track_levels for d in _by_level[lv]]
            st.caption(f'전체 {len(track_all)}개')
            dl_btn('⬇ 계열 선택 과목 전체', track_all, '계열선택과목_성취기준_전체', 'dl_track_all')
        with col_t2:
            sel_track = st.selectbox('계열 선택 과목', track_levels, key='dl_track_sel',
                                     format_func=lambda l: f'{l}  ({_lv_counts.get(l,0)}개)')
            track_idx = _sub_index(_by_level.get(sel_track, []))
            sel_tr_sub = st.selectbox('과목', sorted(track_idx), key='dl_track_sub',
                                      format_func=lambda s: f'{s}  ({len(track_idx[s])}개)')
            tr_items = track_idx.get(sel_tr_sub, [])
            dl_btn(f'⬇ {sel_tr_sub} ({len(tr_items)}개)', tr_items,
                   f'{sel_track}_{sel_tr_sub}_성취기준', 'dl_track_item')

    # ── 전문교과 (인덱스 사용) ───────────────────────────────
    with dl_tab_voc:
        voc_levels = [l for l in LEVEL_ORDER[7:] if l in _by_level]
        col_v1, col_v2 = st.columns([1, 2])
        with col_v1:
            voc_all = [d for lv in voc_levels for d in _by_level[lv]]
            st.caption(f'전체 {len(voc_all):,}개')
            dl_btn('⬇ 전문교과 전체', voc_all, '전문교과_성취기준_전체', 'dl_voc_all')
        with col_v2:
            sel_voc = st.selectbox('전문교과', voc_levels, key='dl_voc_sel',
                                   format_func=lambda l: f'{l}  ({_lv_counts.get(l,0):,}개)')
            voc_idx = _sub_index(_by_level.get(sel_voc, []))
            sel_voc_sub = st.selectbox('과목', sorted(voc_idx), key='dl_voc_sub',
                                       format_func=lambda s: f'{s}  ({len(voc_idx[s])}개)')
            voc_items = voc_idx.get(sel_voc_sub, [])
            dl_btn(f'⬇ {sel_voc_sub} ({len(voc_items)}개)', voc_items,
                   f'{sel_voc}_{sel_voc_sub}_성취기준', 'dl_voc_item')

# ─────────────────────────────────────────────────────────
#  8. 사이드바: 편집 현황 + 빠른 다운로드
# ─────────────────────────────────────────────────────────
with st.sidebar:
    # ── 빠른 다운로드 ───────────────────────────────────
    st.markdown('### 📥 빠른 다운로드')
    dl_btn('⬇ 중학교 전체', _by_level.get('중학교', []), '중학교_성취기준_전체', 'sb_dl_mid')
    dl_btn('⬇ 고등학교 전체', _by_level.get('고등학교', []), '고등학교_성취기준_전체', 'sb_dl_high')
    dl_btn('⬇ 계열 선택 과목 전체', [d for lv,items in _by_level.items() if LEVEL_GROUP.get(lv)=='계열 선택 과목' for d in items], '계열선택과목_성취기준_전체', 'sb_dl_track')
    dl_btn('⬇ 전문교과 전체', [d for lv,items in _by_level.items() if LEVEL_GROUP.get(lv)=='전문교과' for d in items], '전문교과_성취기준_전체', 'sb_dl_voc')
    dl_btn('⬇ 전체', data, '2022개정_성취기준_전체', 'sb_dl_all')

    st.divider()

    # ── 수정 현황 ────────────────────────────────────────
    st.markdown('### 📝 수정 현황')
    edited_codes = list(st.session_state.edits.keys())
    if edited_codes:
        st.caption(f'총 {len(edited_codes)}건 수정됨')
        for code in edited_codes:
            st.markdown(f'- `{code}`')
        st.divider()
        if st.button('🔄 서버에서 최신 수정본 불러오기'):
            load_edits_remote.clear()
            st.session_state.edits = load_edits_remote()
            st.rerun()
    else:
        st.caption('수정된 성취기준 없음')
